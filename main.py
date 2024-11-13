import customtkinter as ctk
import sqlite3
from tkinter import messagebox
import threading  # Import the threading module
import time  # Import the time module
import paho.mqtt.client as mqtt
import openpyxl
from openpyxl.styles import PatternFill
import networkx as nx
from warehouse_MAP import warehouse_graph
from input import previous_location, start_location as start_point, target_location as target_point
from input import mission, obstacles, import_location as impor, export_location as expor, charger_location as charger
from Path import Path


# MQTT Broker
broker_address = "91.121.93.94"
broker_port = 1883

# MQTT Topics
input_topic = "inagv"
output_topic = "outagv"

# Global variable to hold the confirmation response
confirmation_response = None

# Callback functions for MQTT events
def on_connect(client, userdata, flags, rc):
    print("Connected to MQTT Broker with result code " + str(rc))
    client.subscribe(input_topic)

def on_message(client, userdata, msg):
    global confirmation_response
    confirmation_response = msg.payload.decode()

# Create MQTT client
client = mqtt.Client()
client.on_connect = on_connect
client.on_message = on_message

# Connect to MQTT Broker
client.connect(broker_address, broker_port, 60)

# Start the MQTT loop
client.loop_start()



def dijkstra_path_planning_MS(graph, start, target, obstacles=None, previous_location=None):
    if obstacles is None:
        obstacles = []

    # Create a copy of the graph to not affect the original graph
    graph_copy = graph.copy()

    # Remove obstacles from the copied graph
    graph_copy.remove_nodes_from(obstacles)

    # Use Dijkstra's algorithm on the modified graph
    path = nx.shortest_path(graph_copy, source=start, target=target, weight='rfid')

    rfid_tags = [graph.nodes[node]['rfid'] for node in path]
    directions = []

    for i, current_node in enumerate(path[:-1]):
        next_node = path[i + 1]
        prev_node = path[i - 1] if i > 0 else previous_location

        if prev_node is not None and next_node == prev_node:
            directions.append(f"180")
        else:
            prev_pos = graph.nodes[prev_node]['pos'] if prev_node is not None else None
            current_pos = graph.nodes[current_node]['pos']

            if prev_pos is not None:
                prev_vector = (prev_pos[0] - current_pos[0], prev_pos[1] - current_pos[1])
                next_vector = (graph.nodes[next_node]['pos'][0] - current_pos[0],
                               graph.nodes[next_node]['pos'][1] - current_pos[1])

                cross_product = prev_vector[0] * next_vector[1] - prev_vector[1] * next_vector[0]

                if cross_product < 0:
                    directions.append(f"L")
                elif cross_product > 0:
                    directions.append(f"R")
                else:
                    directions.append(f"S")
            else:
                directions.append(f"S")

    if graph.degree(target) == 1 and len(path) >= 2 and current_node == path[-2]:
        directions.append("U")

    directions.append(f"F")

    return rfid_tags, directions, path

def dijkstra_path_planning_ES(graph, start, target, obstacles=None, invert_first_turn=False):
    if graph.degree(start) > 1:
        print(f"Warning: The start node {start} has more than one edge connected to it. "
              f"The path planning might not work as expected.")
    if obstacles is None:
        obstacles = []

    graph_copy = graph.copy()
    graph_copy.remove_nodes_from(obstacles)

    path = nx.shortest_path(graph_copy, source=start, target=target, weight='weight')
    rfid_tags = [graph.nodes[node]['rfid'] for node in path]
    directions = []

    for i in range(1, len(path) - 1):
        prev_node, current_node, next_node = path[i - 1], path[i], path[i + 1]
        prev_vector = (graph.nodes[prev_node]['pos'][0] - graph.nodes[current_node]['pos'][0],
                       graph.nodes[prev_node]['pos'][1] - graph.nodes[current_node]['pos'][1])
        next_vector = (graph.nodes[next_node]['pos'][0] - graph.nodes[current_node]['pos'][0],
                       graph.nodes[next_node]['pos'][1] - graph.nodes[current_node]['pos'][1])
        cross_product = prev_vector[0] * next_vector[1] - prev_vector[1] * next_vector[0]
        cross_product *= -1

        if i == 1 and invert_first_turn:
            direction = f"R" if cross_product > 0 else (
                f"L" if cross_product < 0 else f"180")
        else:
            direction = f"L" if cross_product > 0 else (
                f"R" if cross_product < 0 else f"S")

        directions.append(direction)

    if graph.degree(start) == 1:
        directions.insert(0, "D")

    if graph.degree(target) == 1 and path[-2] == target:
        directions.append("U")

    directions.insert(0, f"B")
    directions.append(f"F")

    return rfid_tags, directions, path

def save_path_and_directions_to_files(path_output_file, directions_output_file, directions, path):
    with open(path_output_file, 'w') as path_file:
        path_file.write(f"Path= {path}\n")

    with open(directions_output_file, 'w') as directions_file:
        directions_file.write("Directions=\n")
        directions_file.write('\n'.join(directions))

def update_previous_location(input_point):

    def find_previous_point(path, input_point):
        index = path.index(input_point)
        if index == 0:
            return None  # There is no previous point for the first point in the path
        else:
            return path[index - 1]

    if input_point in Path:
        previous_point = find_previous_point(Path, input_point)
        if previous_point:
            with open("input.py", "r") as file:
                lines = file.readlines()
            with open("input.py", "w") as file:
                for line in lines:
                    if line.startswith("previous_location"):
                        file.write(f"previous_location = '{previous_point}'\n")
                    else:
                        file.write(line)
            return f"The previous point to {input_point} is {previous_point}"
        else:
            return f"{input_point} is the first point in the path and has no previous point."
    else:
        return f"{input_point} is not in the path."

def save_path_and_directions_to_files(path_output_file, directions_output_file, directions, path):
    with open(path_output_file, 'w') as path_file:
        path_file.write(f"Path= {path}\n")

    with open(directions_output_file, 'w') as directions_file:
        directions_file.write("Directions=\n")
        for direction in directions:
            directions_file.write(f"{direction}\n")

def read_tasks_from_excel(file_path):
    tasks = []
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        if len(row) < 2:
            continue  # Skip rows with less than two elements
        elif len(row) == 2:
            tasks.append((row[0], row[1], 'Incomplete'))  # Append default value for the 3rd column
        else:
            tasks.append((row[0], row[1], row[2]))
    return tasks

def update_excel_sheet(file_path, row_index):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    sheet.cell(row=row_index, column=3, value="COMPLETED")
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=row_index, column=col).fill = PatternFill(start_color="c6efce", end_color="2a734a",fill_type="solid")
    wb.save(file_path)

def run_appropriate_code(start_location, target_location):
    graph, _ = warehouse_graph()
    # Check the degree of the location node
    location_degree = graph.degree(start_location)
    # Run the appropriate code based on the location degree
    if location_degree == 1:
        print("run_edge_start_code")
        branched_warehouse_graph, branched_pos = warehouse_graph()
        nx.set_node_attributes(branched_warehouse_graph, branched_pos, 'pos')

        rfid_path_branched, directions_branched, numeric_path_branched = dijkstra_path_planning_ES(
            branched_warehouse_graph, start_location, target_location, obstacles, invert_first_turn=True
        )

        path_output_file = 'Path.py'
        directions_output_file = 'Directions.txt'

        save_path_and_directions_to_files(path_output_file, directions_output_file, directions_branched,
                                          numeric_path_branched)
        return directions_branched  # Return directions_branched

    else:
        print("run_middle_start_code")

        branched_warehouse_graph, pos = warehouse_graph()
        nx.set_node_attributes(branched_warehouse_graph, pos, 'pos')

        rfid_path_branched, directions_branched, numeric_path_branched = dijkstra_path_planning_MS(
            branched_warehouse_graph, start_location, target_location, obstacles, previous_location
        )
        path_output_file = 'Path.py'
        directions_output_file = 'Directions.txt'
        save_path_and_directions_to_files(path_output_file, directions_output_file, directions_branched,
                                          numeric_path_branched)
        return directions_branched  # Return directions_branched

def start_AGV():
    global confirmation_response  # Declare it global here

    tasks = read_tasks_from_excel('tasks.xlsx')
    final_point=None

    for i, task in enumerate(tasks, start=1):
        target_point, mission, status = task
        if status != "COMPLETED":
            if mission == 'import':

                final_point = target_point
                update_previous_location(start_point)
                directions_branched = run_appropriate_code(start_point, impor)
                for direction in directions_branched:
                    client.publish(output_topic, direction)
                    confirmation_response = None  # Reset the confirmation response
                    while confirmation_response is None:
                        continue
                    if confirmation_response.lower() != 'y':
                        client.publish(output_topic, "Confirmation not received. Exiting.")
                        return

                update_previous_location(impor)
                directions_branched = run_appropriate_code(impor, target_point)
                for direction in directions_branched:
                    client.publish(output_topic, direction)
                    confirmation_response = None  # Reset the confirmation response
                    while confirmation_response is None:
                        continue
                    if confirmation_response.lower() != 'y':
                        client.publish(output_topic, "Confirmation not received. Exiting.")
                        return

            elif mission == 'export':

                final_point = expor
                update_previous_location(start_point)
                directions_branched = run_appropriate_code(start_point, target_point)
                for direction in directions_branched:
                    client.publish(output_topic, direction)
                    confirmation_response = None  # Reset the confirmation response
                    while confirmation_response is None:
                        continue
                    if confirmation_response.lower() != 'y':
                        client.publish(output_topic, "Confirmation not received. Exiting.")
                        return

                update_previous_location(target_point)
                directions_branched = run_appropriate_code(target_point, expor)
                for direction in directions_branched:
                    client.publish(output_topic, direction)
                    confirmation_response = None  # Reset the confirmation response
                    while confirmation_response is None:
                        continue
                    if confirmation_response.lower() != 'y':
                        client.publish(output_topic, "Confirmation not received. Exiting.")
                        return

            update_excel_sheet('tasks.xlsx', i)
            client.publish(output_topic, f"Task to {target_point} marked as completed.")

    client.publish(output_topic, "All tasks completed!")
    update_previous_location(expor)
    directions_branched = run_appropriate_code(final_point, charger)
    for direction in directions_branched:
        client.publish(output_topic, direction)
        confirmation_response = None  # Reset the confirmation response
        while confirmation_response is None:
            continue
        if confirmation_response.lower() != 'y':
            client.publish(output_topic, "Confirmation not received. Exiting.")
            return

def start_AGV_in_thread():
    threading.Thread(target=start_AGV).start()

class PlanStock(ctk.CTkCanvas):
    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)
        
        self.stations = {}
        self.lines = []
        self.stations1 = {}
        self.blinking_stations = set()
        

    def add_station(self, name, x, y, radius=15):  # Augmenter le rayon pour rendre les cercles plus grands
        station = self.create_oval(x-radius, y-radius, x+radius, y+radius, fill='black')
        #self.create_text(x, y-radius-8, text=name, font=("Arial", 12, "bold"))  # Déplacer le texte légèrement vers le haut et augmenter la taille de la police
        self.stations[name] = (x, y)
        self.stations1[name] = station
        
        
    

    def add_line(self, start, end,color='black'):
        x1, y1 = self.stations[start]
        x2, y2 = self.stations[end]
        line = self.create_line(x1, y1, x2, y2, fill=color, width=4)
        self.lines.append(line)


    def add_line_2(self, start, end,color='black'):
        x1, y1 = self.stations[start]
        x2, y2 = self.stations[end]
        if y1>y2 :
            line = self.create_line(x1, y1, x2, y2+20, fill=color, width=4)
            self.lines.append(line)
        elif  y1<y2:
            line = self.create_line(x1, y1, x2, y2-20, fill=color, width=4)
            self.lines.append(line)
        if x2>x1:
            line = self.create_line(x1, y1, x2-20, y2, fill=color, width=4)
            self.lines.append(line)
        elif  x1>x2:
            line = self.create_line(x1, y1, x2+20, y2, fill=color, width=4)
            self.lines.append(line)

    def add_square(self, name, x, y, size=20, color='grey'):
        square = self.create_rectangle(x-size, y-size, x+size, y+size, fill=color)
        self.create_text(x, y, text=name, font=("Arial", 12, "bold"))
        self.stations[name] = (x, y)
        self.stations1[name] = square
    
    def highlight_station(self, name):
        if name not in self.blinking_stations:
            station_id = self.stations1[name]
            self.itemconfig(station_id, fill='grey')
            self.after(550, self.change_color, station_id, 'red',name)  # Change color to red after 500 milliseconds
            self.blinking_stations.add(name)

    def change_color(self, station_id, color, name):
        if name not in emplacements:
            self.blinking_stations.remove(name)  # Stop the blinking cycle
            return
        self.itemconfig(station_id, fill=color)
        if color == 'red':
            self.after(550, self.change_color, station_id, 'grey', name)
        else:
            self.after(550, self.change_color, station_id, 'red', name)

    def update_stations(self):
        global emplacements
        conn = sqlite3.connect('products_base.db')
        cursor = conn.cursor()
        cursor.execute('''SELECT emplacement FROM products_data WHERE reference IS NULL AND categorie IS NULL OR reference = '' AND categorie = '';''')
        results = cursor.fetchall()
        emplacements = [row[0] for row in results]
        for name in emplacements:
            self.highlight_station(name)
            
        self.after(1100, self.update_stations)

    def addproduct_highlt(self, name,color):
        if name in self.stations1:
            station_id = self.stations1[name]
            self.itemconfig(station_id, fill=color)
    
    def product_available(self):
        
        cursor.execute('''SELECT emplacement FROM products_data WHERE reference IS NOT NULL AND categorie IS NOT NULL OR reference <> '' AND categorie <> '';''')
        result = cursor.fetchall()
        fullstation = [row[0] for row in result]
        for fullname in fullstation:
            station_id = self.stations1[fullname]
            self.itemconfig(station_id, fill='green')
    
    def outproduct_highlt(self, name,color):
        if name in self.stations1:
            station_id = self.stations1[name]
            self.itemconfig(station_id, fill=color)



            




root = ctk.CTk()
root.title('inventory')
root.geometry()

conn = sqlite3.connect('products_base.db')
cursor = conn.cursor()

table_create_query = '''CREATE TABLE IF NOT EXISTS products_data
                        (reference TEXT , categorie TEXT , emplacement TEXT )
                    '''
cursor.execute(table_create_query)

# Add a global flag to check if the product is placed

product_placed = True
######################################
def addproduct():
    global spot
    global product_placed
    global blinking_flag

    # Check if the previous product is placed
    if not product_placed:
        messagebox.showerror('Erreur', "The first product hasn't been placed yet")
        return

    reference = reference_entry.get()
    catagory = catagory_entry.get()

    if reference != '' and catagory != '':
        cursor.execute('''SELECT emplacement FROM products_data WHERE reference IS NULL AND categorie IS NULL OR reference = '' AND categorie = '';''')
        result = cursor.fetchone()

        if result is not None:
            spot = result[0]
            cursor.execute('UPDATE products_data SET reference = ?, categorie = ? WHERE emplacement = ?', [reference, catagory, result[0]])
            conn.commit()
            messagebox.showinfo("Success", "Produit ajouté")

            # Set the product_placed flag to False as the product is not yet placed
            product_placed = False

            # Blinking process
            blinking_flag = True

            def blink_spot():
                while blinking_flag:
                    canvas.addproduct_highlt(spot, 'blue')  # Highlight in blue
                    time.sleep(0.55)
                    canvas.addproduct_highlt(spot, 'green')  # Highlight in green
                    time.sleep(0.55)

                print("Blinking stopped.")

            # Start the blinking process in a separate thread
            threading.Thread(target=blink_spot).start()

            # Wait for user input in the terminal
            user_input = input("Enter 'ok' to stop blinking: ")
            if user_input.lower() == "ok":
                blinking_flag = False  # Set flag to False to stop blinking
                product_placed = True  # Set the product_placed flag to True as the product is now placed
                messagebox.showinfo("Success", "Produit est placé")

        else:
            messagebox.showerror('Message', "Tous les emplacements sont occupés")
    else:
        messagebox.showerror('Erreur', "Tous les champs sont requis")


product_out=True

def outproduct():
    global blinking_full
    global fullspot
    global product_out
    if not product_out:
        messagebox.showerror('Erreur', "The first product hasn't been out yet")
        return

    referenceoutput=reference_out.get()
    if referenceoutput !='':
        cursor.execute('''SELECT emplacement FROM products_data WHERE reference = ?;''',[referenceoutput])
        result = cursor.fetchone()
        
        if result is not None:
            fullspot = result[0]
            
            product_out = False
            blinking_full = True
            def blinkfull_spot():
                while blinking_full:
                    canvas.outproduct_highlt(fullspot, 'green')  # Highlight in blue
                    time.sleep(0.55)
                    canvas.outproduct_highlt(fullspot, 'orange')  # Highlight in green
                    time.sleep(0.55)

                print("Blinking stopped.")

            # Start the blinking process in a separate thread
            threading.Thread(target=blinkfull_spot).start()
            user_input = input("Enter 'done' to stop blinking: ")
            if user_input.lower() == "done":
                blinking_full = False
                product_out = True 
                messagebox.showinfo("Success", "product out") 
                cursor.execute('UPDATE products_data SET reference = NULL, categorie = NULL WHERE emplacement = ?', [fullspot])
                conn.commit() 
                
        
        else:
            messagebox.showerror('Erreur','Ce produit n\'est pas dans le système')
    else:
        messagebox.showerror('erreur',"Veuillez saisir la référence du produit à sortir")





    



frame = ctk.CTkFrame(root)

frame.pack()

#entry_out_DB_frame

entry_out_DB_frame=ctk.CTkFrame(frame)

#entré produit
Product_entry_frame = ctk.CTkFrame(entry_out_DB_frame)

entry_product_label= ctk.CTkLabel(Product_entry_frame,text="stockage produit")
reference_entry = ctk.CTkEntry(Product_entry_frame,placeholder_text='Référence')
catagory_entry = ctk.CTkEntry(Product_entry_frame,placeholder_text='catégorie')
entry_button = ctk.CTkButton(Product_entry_frame, text="Ajouté",command=addproduct)
#sortir produit
Product_out_frame = ctk.CTkFrame(entry_out_DB_frame)

out_product_label= ctk.CTkLabel(Product_out_frame,text="sortir de produit")
reference_out = ctk.CTkEntry(Product_out_frame,placeholder_text='Référence')
out_button = ctk.CTkButton(Product_out_frame, text="sortir",command=outproduct)

#affichage de la base de donné
base_de_donner_frame= ctk.CTkFrame(entry_out_DB_frame)
base_donné_label= ctk.CTkLabel(base_de_donner_frame,text="Base des donnés")
affiBD_button = ctk.CTkButton(base_de_donner_frame, text="Afficher la BD")
suppBD_button = ctk.CTkButton(base_de_donner_frame, text="supprimer un article")

#message et map frame
map_frame = ctk.CTkFrame(frame)

scrollable_frame = ctk.CTkScrollableFrame(frame ,label_text="message output",width=400)



action_frame = ctk.CTkFrame(frame)

efface_button= ctk.CTkButton(action_frame,text="supprimer un article")
logiciel_button= ctk.CTkButton(action_frame,text="logiciel de gestion")
robot1_button = ctk.CTkButton(action_frame, text="Start AGV", command=start_AGV_in_thread)

#affichage de layout
entry_out_DB_frame.grid(row=0,column=0,columnspan=2,padx=20,pady=5,sticky='news')
scrollable_frame.grid(row=1,column=0,padx=20,pady=5, sticky='nswe')
map_frame.grid(row=1,column=1,padx=20,pady=5, sticky='nsew')
action_frame.grid(row=2,column=0,columnspan=2,sticky="news",padx=20,pady=5)


Product_out_frame.pack(expand = True,side=ctk.LEFT,fill=ctk.BOTH,padx=10,pady=10)
Product_entry_frame.pack(expand = True,side=ctk.LEFT,fill=ctk.BOTH,padx=10,pady=10)
base_de_donner_frame.pack(expand = True,side=ctk.LEFT,fill=ctk.BOTH,padx=10,pady=10)
#widget affichage

#entré
entry_product_label.pack(pady =10)
reference_entry.pack(padx=10,pady=10)
catagory_entry.pack(padx=10,pady=10)
entry_button.pack(padx=10,pady=10)
#sortir
out_product_label.pack(pady=10)
reference_out.pack(padx=10,pady=10)
out_button.pack(padx=10,pady=10)
#baseD

base_donné_label.pack(pady=10)
affiBD_button.pack(padx=10,pady=10)
suppBD_button.pack(padx=10,pady=10)




#logiciel et effacer

efface_button.pack(side=ctk.LEFT,padx=20)
logiciel_button.pack(side=ctk.RIGHT,padx=20)
robot1_button.pack(side=ctk.RIGHT,padx=20)

canvas = PlanStock(map_frame, width=620, height=590, highlightthickness=0)
canvas.pack(padx=5,pady=5)

    # Ajout des stations avec des cercles plus grands

canvas.add_station("A1", 100, 100, radius=10)
canvas.add_station("A2", 170, 100, radius=10)
canvas.add_station("A3", 240, 100, radius=10)
canvas.add_station("A4", 310, 100, radius=10)
canvas.add_station("A5", 380, 100, radius=10)
canvas.add_station("A6", 450, 100, radius=10)
canvas.add_station("A7", 520, 100, radius=10)
canvas.add_station("A8", 100, 200, radius=10)
canvas.add_station("A9", 310, 200, radius=10)
canvas.add_station("A10", 520, 200,radius=10)   
canvas.add_station("A11", 100, 300, radius=10)
canvas.add_station("A12", 310, 300,radius=10)
canvas.add_station("A13", 520, 300,radius=10) 
canvas.add_station("A14", 100, 400, radius=10)
canvas.add_station("A15", 310, 400,radius=10)
canvas.add_station("A16", 520, 400,radius=10) 
canvas.add_station("A17", 100, 500, radius=10)
canvas.add_station("A18", 310, 500,radius=10)
canvas.add_station("A19", 520, 500,radius=10) 

    #placer les rectangles
square_liste_1 = ["FB", "FC", "FD", "FE", "FF", "FG", "FH"]
i=100
for name_st_1 in square_liste_1:
    canvas.add_square(name_st_1, i, 50)
    i+=70
square_liste_2 = ["FA","EA","DA","CA","BA","FI","EI","DI","CI","BI",
                      "EC","CC","ED","CD","EF","CF","EG","CG","AE"] 
j=j1=100  
  
j2=j3=j4=j5=200  
for name_st_2 in square_liste_2:
    if  j<600 :
        canvas.add_square(name_st_2,50,j)
        j+=100
    elif j1<= 500 :
        canvas.add_square(name_st_2,570,j1)
        j1+=100
    elif j2<= 400:
        canvas.add_square(name_st_2,150,j2)
        j2+=200
    elif j3<= 400:
        canvas.add_square(name_st_2,260,j3)
        j3+=200
    elif j4<= 400:
        canvas.add_square(name_st_2,360,j4)
        j4+=200      
    elif j5<= 400:
        canvas.add_square(name_st_2,470,j5)
        j5+=200
    else :
        canvas.add_square(name_st_2,310,550)


    # Ajout des lignes
for i in range (1, 7):
    canvas.add_line("A"+str(i),"A"+str(i+1))
canvas.add_line("A1","A8")
for i in range (8, 17, 3):
    canvas.add_line("A"+str(i),"A"+str(i+3))
canvas.add_line("A4","A9")
for i in range (9, 18, 3):
    canvas.add_line("A"+str(i),"A"+str(i+3))
canvas.add_line("A7","A10")
for i in range (10, 19, 3):
    canvas.add_line("A"+str(i),"A"+str(i+3))
for i in range (17, 19):
    canvas.add_line("A"+str(i),"A"+str(i+1))
for i in range (11, 13):
    canvas.add_line("A"+str(i),"A"+str(i+1))
    
    #ajoute de ligne pour les carreaux
y=0
temp_list = ["A1","A2","A3","A4","A5","A6","A7","A1","A8","A11",
                      "A14","A17","A7","A10","A13","A16","A19","A8"
                      ,"A14","A9","A15","A9","A15","A10","A16","A18"]
stations_list_keys = list(canvas.stations.keys())
start_square_key="FB"
key_index = stations_list_keys.index(start_square_key)
    
for cle in stations_list_keys[key_index:]:


    canvas.add_line_2(temp_list[y],cle,"black")
    y+=1


canvas.update_stations()

canvas.product_available()

root.mainloop()